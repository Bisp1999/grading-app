from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, session, current_app, send_file
from flask_login import login_user, logout_user, login_required, current_user, LoginManager
from flask_babel import gettext as _
from werkzeug.security import check_password_hash, generate_password_hash
from . import db, login_manager
from .models import Teacher, School, Classroom, Student, SetupWizardData, Test, Grade, ClassroomLayout
from .forms import LoginForm, RegistrationForm
import json
import math
from datetime import datetime, date
import logging
import os
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
import smtplib
from email.mime.text import MIMEText

main = Blueprint('main', __name__)

@main.route('/set_language/<language>')
def set_language(language=None):
    """Set user's preferred language"""
    if language and language in ['en', 'fr']:
        session['language'] = language
        session.permanent = True  # Make session permanent to ensure it persists
        
        # If user is logged in, save to their profile
        if current_user.is_authenticated:
            current_user.preferred_language = language
            db.session.commit()
        
    
    # If coming from language selector, redirect to appropriate next step
    if request.referrer and 'language_selector' in request.referrer:
        if current_user.is_authenticated:
            response = redirect(url_for('main.dashboard'))
        else:
            response = redirect(url_for('main.login'))
    else:
        # Redirect back to referrer or home page
        response = redirect(request.referrer or url_for('main.index'))
    
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@main.route('/language_selector')
def language_selector():
    """Show language selection page for non-authenticated users"""
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    
    return render_template('language_selector.html')

@main.route('/toggle_language')
@login_required
def toggle_language():
    """Toggle language between English and French for testing purposes"""
    current_lang = current_user.preferred_language if current_user.preferred_language else 'en'
    new_lang = 'fr' if current_lang == 'en' else 'en'
    
    # Update user's preferred language
    current_user.preferred_language = new_lang
    db.session.commit()
    
    # Also update session
    session['language'] = new_lang
    
    flash(_('Language changed successfully!'), 'success')
    return redirect(request.referrer or url_for('main.dashboard'))

def extract_grade_from_classroom_name(classroom_name):
    """Extract grade from classroom name (e.g., '101' -> 'Grade 1', '201' -> 'Grade 2')"""
    import re
    
    # Try to extract grade from classroom name patterns
    # Pattern 1: "101", "102" -> Grade 1
    # Pattern 2: "201", "202" -> Grade 2
    # Pattern 3: "Grade 1", "Grade 2" -> Grade 1, Grade 2
    # Pattern 4: "1A", "2B" -> Grade 1, Grade 2
    
    if not classroom_name:
        return 'Unknown Grade'
    
    # Check if it already contains "Grade"
    if 'grade' in classroom_name.lower():
        return classroom_name
    
    # Try to extract number patterns
    # Look for patterns like 101, 201, 301 (first digit is grade)
    match = re.match(r'^([1-9])\d+$', classroom_name)
    if match:
        grade_num = match.group(1)
        return f'Grade {grade_num}'
    
    # Look for patterns like 1A, 2B, 3C (first character is grade)
    match = re.match(r'^([1-9])[A-Za-z]+$', classroom_name)
    if match:
        grade_num = match.group(1)
        return f'Grade {grade_num}'
    
    # Look for any number at the beginning
    match = re.match(r'^([1-9])', classroom_name)
    if match:
        grade_num = match.group(1)
        return f'Grade {grade_num}'
    
    # If no pattern matches, return the classroom name as-is
    return classroom_name

@main.route('/')
def index():
    # If user is not authenticated and hasn't selected a language, show language selector
    if not current_user.is_authenticated and 'language' not in session:
        return redirect(url_for('main.language_selector'))
    return redirect(url_for('main.dashboard'))

# ---------------------- Password Reset Flow ----------------------

def _get_serializer():
    from flask import current_app
    secret = current_app.config['SECRET_KEY']
    return URLSafeTimedSerializer(secret_key=secret, salt='password-reset')

def _send_reset_email(to_email: str, reset_url: str):
    """Send a password reset email. If SMTP isn't configured, flash the link as a fallback."""
    from flask import current_app
    mail_server = current_app.config.get('MAIL_SERVER') or os.environ.get('MAIL_SERVER')
    mail_port = int(current_app.config.get('MAIL_PORT') or os.environ.get('MAIL_PORT') or 0)
    mail_username = current_app.config.get('MAIL_USERNAME') or os.environ.get('MAIL_USERNAME')
    mail_password = current_app.config.get('MAIL_PASSWORD') or os.environ.get('MAIL_PASSWORD')
    mail_use_tls = (str(current_app.config.get('MAIL_USE_TLS') or os.environ.get('MAIL_USE_TLS') or 'false')).lower() == 'true'
    mail_use_ssl = (str(current_app.config.get('MAIL_USE_SSL') or os.environ.get('MAIL_USE_SSL') or 'false')).lower() == 'true'

    subject = 'Password Reset Instructions'
    body = f"Click the link to reset your password: {reset_url}\nIf you did not request this, please ignore."

    # No SMTP configured: show the link as a development fallback
    if not mail_server:
        flash(_('Password reset link (development): %(link)s', link=reset_url), 'info')
        return

    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = mail_username or 'no-reply@example.com'
    msg['To'] = to_email

    try:
        if mail_use_ssl:
            with smtplib.SMTP_SSL(mail_server, mail_port or 465) as server:
                if mail_username and mail_password:
                    server.login(mail_username, mail_password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(mail_server, mail_port or 587) as server:
                if mail_use_tls:
                    server.starttls()
                if mail_username and mail_password:
                    server.login(mail_username, mail_password)
                server.send_message(msg)
    except Exception as e:
        logging.exception('Failed to send reset email')
        flash(_('Could not send reset email. Please contact support.'), 'danger')

@main.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        if not email:
            flash(_('Please enter your email address.'), 'warning')
            return redirect(url_for('main.forgot_password'))

        user = Teacher.query.filter_by(email=email).first()
        
        # Always show the same message to avoid revealing which emails are registered
        if user:
            s = _get_serializer()
            token = s.dumps({'email': email})
            reset_url = url_for('main.reset_password', token=token, _external=True)
            _send_reset_email(email, reset_url)

        flash(_('If an account exists for that email, a reset link has been sent.'), 'info')
        return redirect(url_for('main.login'))

    return render_template('login.html', mode='forgot')

@main.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    s = _get_serializer()
    try:
        data = s.loads(token, max_age=3600)  # 1 hour
        email = data.get('email')
    except SignatureExpired:
        flash(_('The reset link has expired. Please request a new one.'), 'danger')
        return redirect(url_for('main.forgot_password'))
    except BadSignature:
        flash(_('Invalid reset link.'), 'danger')
        return redirect(url_for('main.forgot_password'))

    if request.method == 'POST':
        password = request.form.get('password') or ''
        confirm = request.form.get('confirm_password') or ''
        if len(password) < 6:
            flash(_('Password must be at least 6 characters.'), 'warning')
            return redirect(request.url)
        if password != confirm:
            flash(_('Passwords do not match.'), 'warning')
            return redirect(request.url)

        user = Teacher.query.filter_by(email=email).first()
        if not user:
            flash(_('Account not found.'), 'danger')
            return redirect(url_for('main.forgot_password'))

        user.password_hash = generate_password_hash(password)
        db.session.commit()
        flash(_('Your password has been updated. Please log in.'), 'success')
        return redirect(url_for('main.login'))

    return render_template('login.html', mode='reset', token=token)

@main.route('/setup_wizard')
@login_required
def setup_wizard():
    return render_template('setup_wizard.html')

@main.route('/student_wizard')
@login_required
def student_wizard():
    return render_template('student_wizard.html')

@main.route('/student_tab')
@login_required
def student_tab():
    """Student Tab page - view individual student details and grades"""
    from .models import SetupWizardData, School, Classroom, Student
    
    # Get Setup Wizard data for the current user
    wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
    
    if not wizard_data:
        flash('Please complete the Setup Wizard first.', 'warning')
        return redirect(url_for('main.setup_wizard'))
    
    # Get teacher type from Setup Wizard data
    teacher_type = wizard_data.teacher_type if wizard_data.teacher_type else 'homeroom'
    
    # Get classroom data
    all_classrooms = []
    classrooms_by_school = {}
    
    for school in School.query.filter_by(teacher_id=current_user.id).all():
        school_classrooms = Classroom.query.filter_by(school_id=school.id).all()
        classrooms_by_school[school.name] = school_classrooms
        all_classrooms.extend(school_classrooms)
    
    # Organize data based on teacher type
    if teacher_type == 'specialist':
        classrooms_by_grade = {}
        for classroom in all_classrooms:
            if '(' in classroom.name and ')' in classroom.name:
                parts = classroom.name.split(' (')
                classroom_name = parts[0]
                grade = parts[1].rstrip(')')
            else:
                classroom_name = classroom.name
                grade = extract_grade_from_classroom_name(classroom.name)
            
            if grade not in classrooms_by_grade:
                classrooms_by_grade[grade] = []
            classrooms_by_grade[grade].append({
                'id': classroom.id,
                'name': classroom_name,
                'full_name': classroom.name
            })
        
        grades = sorted(classrooms_by_grade.keys())
    else:
        grades = []
        classrooms_by_grade = {}
    
    return render_template('student_tab.html',
                         teacher_type=teacher_type,
                         grades=grades,
                         classrooms_by_grade=classrooms_by_grade,
                         all_classrooms=all_classrooms,
                         show_global_filters=True)

@main.route('/input_grades', methods=['GET', 'POST'])
@login_required
def input_grades():
    """Input test grades page"""
    from .models import SetupWizardData, Test, Grade, Student, Classroom, School
    
    # Get Setup Wizard data for the current user
    wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
    
    if not wizard_data:
        flash('Please complete the Setup Wizard first.', 'warning')
        return redirect(url_for('main.setup_wizard'))
    
    # Get teacher type from Setup Wizard data (not from classroom count)
    teacher_type = wizard_data.teacher_type if wizard_data.teacher_type else 'homeroom'
    
    # Get classroom data
    all_classrooms = []
    for school in School.query.filter_by(teacher_id=current_user.id).all():
        school_classrooms = Classroom.query.filter_by(school_id=school.id).all()
        all_classrooms.extend(school_classrooms)
    
    # Get all tests for this teacher
    tests = Test.query.filter_by(teacher_id=current_user.id).order_by(Test.test_date.desc()).all()
    
    if not tests:
        flash('Please create some tests first before inputting grades.', 'info')
        return redirect(url_for('main.create_tests'))
    
    # Calculate grades completion status for each test
    for test in tests:
        # Get students for this test based on teacher type and test details
        if teacher_type == 'specialist':
            # For specialist teachers, get students from the specific class.
            # Do NOT depend on parsing/formatting Classroom.name like "Class (Grade X)", since
            # grade labels are teacher-defined (e.g., "5" vs "Grade 5").

            def _normalize_grade_label(value):
                if value is None:
                    return ''
                s = str(value).strip()
                if not s:
                    return ''
                lower = s.lower()
                if lower.startswith('grade '):
                    s = s[6:].strip()
                return s

            normalized_test_grade = _normalize_grade_label(test.grade)
            target_classroom = None
            class_name_candidates = []

            logger = current_app.logger
            input_grades_debug = (os.environ.get('INPUT_GRADES_DEBUG', '') or '').strip().lower() in ('1', 'true', 'yes', 'on')
            log_debug = input_grades_debug and logger.isEnabledFor(logging.DEBUG)
            if log_debug:
                logger.debug(
                    "INPUT_GRADES: Test %s - class_name='%s', grade='%s'",
                    test.id,
                    test.class_name,
                    test.grade,
                )

            for classroom in all_classrooms:
                if '(' in classroom.name and ')' in classroom.name:
                    parts = classroom.name.split(' (')
                    classroom_name = parts[0]
                    grade = parts[1].rstrip(')')
                else:
                    classroom_name = classroom.name
                    grade = extract_grade_from_classroom_name(classroom.name)

                class_name_matches = (
                    test.class_name == classroom_name or
                    test.class_name == classroom.name
                )
                if not class_name_matches:
                    continue

                normalized_classroom_grade = _normalize_grade_label(grade)
                grade_matches = (not normalized_test_grade) or (normalized_test_grade == normalized_classroom_grade)
                class_name_candidates.append((classroom, grade_matches))

            if class_name_candidates:
                matching_grade_candidates = [c for c in class_name_candidates if c[1]]
                if matching_grade_candidates:
                    target_classroom = matching_grade_candidates[0][0]
                else:
                    target_classroom = class_name_candidates[0][0]

            if target_classroom:
                students = Student.query.filter_by(classroom_id=target_classroom.id).all()
            else:
                students = []

            if log_debug:
                logger.debug("INPUT_GRADES: Found %s students for test %s", len(students), test.id)
        else:
            # For homeroom teachers, get students from all classes
            students = Student.query.join(Classroom).join(School).filter(
                School.teacher_id == current_user.id
            ).all()
            logger = current_app.logger
            input_grades_debug = (os.environ.get('INPUT_GRADES_DEBUG', '') or '').strip().lower() in ('1', 'true', 'yes', 'on')
            log_debug = input_grades_debug and logger.isEnabledFor(logging.DEBUG)
            if log_debug:
                logger.debug(
                    "INPUT_GRADES: Test %s - Homeroom teacher, found %s students",
                    test.id,
                    len(students),
                )
        
        # Count how many of the relevant students have grades for this test
        student_ids = [student.id for student in students]
        
        # Get all grade records for this test and these students
        all_grades = Grade.query.filter_by(test_id=test.id).filter(
            Grade.student_id.in_(student_ids)
        ).all()
        
        # Count students with either a grade OR marked as absent
        # A student is "graded" if they have a grade record with either:
        # 1. A numeric grade (grade is not None), OR
        # 2. Marked as absent (absent = True)
        graded_count = sum(1 for grade in all_grades if (grade.grade is not None or grade.absent))
        
        # Count absent students (only those with Grade records marked as absent)
        absent_count = sum(1 for grade in all_grades if grade.absent)

        logger = current_app.logger
        input_grades_debug = (os.environ.get('INPUT_GRADES_DEBUG', '') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        log_debug = input_grades_debug and logger.isEnabledFor(logging.DEBUG)
        if log_debug:
            logger.debug(
                "INPUT_GRADES: Test %s (%s): %s students, %s grade records, %s graded, %s absent",
                test.id,
                test.test_name,
                len(students),
                len(all_grades),
                graded_count,
                absent_count,
            )
        
        # Set grades_complete attribute
        # Grading is complete when ALL students have been graded (have grade OR absent)
        test.grades_complete = (graded_count == len(students) and len(students) > 0)
        if log_debug:
            logger.debug(
                "INPUT_GRADES: Test %s grades_complete: %s (graded_count=%s, len(students)=%s)",
                test.id,
                test.grades_complete,
                graded_count,
                len(students),
            )
        
        # Set has_absent_students attribute
        # Check if there are any students marked as absent for this test
        test.has_absent_students = absent_count > 0
    
    # Sort tests by: Grading Completed (ascending), Absent Students (descending), Date (ascending)
    # Priority: incomplete grading first, then tests with absent students first
    tests.sort(key=lambda t: (
        t.grades_complete,      # False (not completed) comes before True (completed)
        not t.has_absent_students,  # False (has absent) comes before True (no absent)
        t.test_date  # Date ascending
    ))
    
    # Parse wizard data
    competencies = json.loads(wizard_data.competencies) if wizard_data.competencies else []
    semesters = [f"Semester {i}" for i in range(1, wizard_data.num_semesters + 1)] if wizard_data.num_semesters else []
    
    # Organize data based on teacher type
    if teacher_type == 'specialist':
        classrooms_by_grade = {}

        # Prefer the authoritative setup wizard data, which preserves the exact grade labels
        # the teacher defined (e.g. "V", "Cinq", "My Favorite Grade").
        wizard_classrooms = []
        try:
            wizard_classrooms = json.loads(wizard_data.classrooms) if wizard_data.classrooms else []
        except Exception:
            wizard_classrooms = []

        if wizard_classrooms:
            for classroom in wizard_classrooms:
                classroom_name = str(classroom.get('name') or '').strip()
                grade = str(classroom.get('grade') or '').strip()

                if not classroom_name or not grade:
                    continue

                if grade not in classrooms_by_grade:
                    classrooms_by_grade[grade] = []
                classrooms_by_grade[grade].append(classroom_name)
        else:
            # Fallback for older data: parse from Classroom.name
            for classroom in all_classrooms:
                if '(' in classroom.name and ')' in classroom.name:
                    parts = classroom.name.split(' (')
                    classroom_name = parts[0]
                    grade = parts[1].rstrip(')')
                else:
                    classroom_name = classroom.name
                    grade = extract_grade_from_classroom_name(classroom.name)

                if grade not in classrooms_by_grade:
                    classrooms_by_grade[grade] = []
                classrooms_by_grade[grade].append(classroom_name)
        
        # Sort grades and classroom names
        for grade in classrooms_by_grade:
            classrooms_by_grade[grade].sort()
        
        subjects = [wizard_data.subject_name] if wizard_data.subject_name else []
        grades = sorted(classrooms_by_grade.keys())
    else:
        subjects = json.loads(wizard_data.subjects) if wizard_data.subjects else []
        grades = []
        classrooms_by_grade = {}
    
    return render_template('input_grades.html',
                         tests=tests,
                         competencies=competencies,
                         semesters=semesters,
                         subjects=subjects,
                         grades=grades,
                         classrooms_by_grade=classrooms_by_grade,
                         teacher_type=teacher_type,
                         show_global_filters=True)

@main.route('/review_grades')
@login_required
def review_grades():
    """Review and adjust grades page"""
    from .models import SetupWizardData, Test, Student, Grade, Classroom, School
    
    # Get Setup Wizard data for the current user
    wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
    
    if not wizard_data:
        flash('Please complete the Setup Wizard first.', 'warning')
        return redirect(url_for('main.setup_wizard'))
    
    # Determine teacher type and get relevant data
    all_classrooms = []
    for school in School.query.filter_by(teacher_id=current_user.id).all():
        school_classrooms = Classroom.query.filter_by(school_id=school.id).all()
        all_classrooms.extend(school_classrooms)
    
    total_classrooms = len(all_classrooms)
    teacher_type = 'specialist' if total_classrooms > 1 else 'homeroom'
    
    # Parse wizard data
    competencies = json.loads(wizard_data.competencies) if wizard_data.competencies else []
    semesters = [f"Semester {i}" for i in range(1, wizard_data.num_semesters + 1)] if wizard_data.num_semesters else []
    
    # Organize data based on teacher type
    if teacher_type == 'specialist':
        classrooms_by_grade = {}
        for classroom in all_classrooms:
            if '(' in classroom.name and ')' in classroom.name:
                parts = classroom.name.split(' (')
                classroom_name = parts[0]
                grade = parts[1].rstrip(')')
            else:
                classroom_name = classroom.name
                grade = extract_grade_from_classroom_name(classroom.name)
            
            if grade not in classrooms_by_grade:
                classrooms_by_grade[grade] = []
            classrooms_by_grade[grade].append(classroom_name)
        
        # Sort grades and classroom names
        for grade in classrooms_by_grade:
            classrooms_by_grade[grade].sort()
        
        subjects = [wizard_data.subject_name] if wizard_data.subject_name else []
        grades = sorted(classrooms_by_grade.keys())
    else:
        subjects = json.loads(wizard_data.subjects) if wizard_data.subjects else []
        grades = []
        classrooms_by_grade = {}
    
    return render_template('review_grades.html',
                         competencies=competencies,
                         semesters=semesters,
                         subjects=subjects,
                         grades=grades,
                         classrooms_by_grade=classrooms_by_grade,
                         teacher_type=teacher_type,
                         show_global_filters=True)

@main.route('/create_tests', methods=['GET', 'POST'])
@login_required
def create_tests():
    from .models import SetupWizardData, Test
    
    # Get Setup Wizard data for the current user
    wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
    
    if not wizard_data:
        flash('Please complete the Setup Wizard first.', 'warning')
        return redirect(url_for('main.setup_wizard'))
    
    # Parse wizard data
    competencies = json.loads(wizard_data.competencies) if wizard_data.competencies else []
    semesters = [f"Semester {i}" for i in range(1, wizard_data.num_semesters + 1)] if wizard_data.num_semesters else []
    
    # Get teacher type from Setup Wizard data (not from classroom count)
    teacher_type = wizard_data.teacher_type if wizard_data.teacher_type else 'homeroom'
    
    # Get classroom data
    all_classrooms = []
    classrooms_by_school = {}
    
    for school in School.query.filter_by(teacher_id=current_user.id).all():
        school_classrooms = Classroom.query.filter_by(school_id=school.id).all()
        classrooms_by_school[school.name] = school_classrooms
        all_classrooms.extend(school_classrooms)
    
    # Organize data based on teacher type
    if teacher_type == 'specialist':
        classrooms_by_grade = {}
        for classroom in all_classrooms:
            if '(' in classroom.name and ')' in classroom.name:
                parts = classroom.name.split(' (')
                classroom_name = parts[0]
                grade = parts[1].rstrip(')')
            else:
                classroom_name = classroom.name
                grade = extract_grade_from_classroom_name(classroom.name)
            
            if grade not in classrooms_by_grade:
                classrooms_by_grade[grade] = []
            classrooms_by_grade[grade].append(classroom_name)
        
        # Sort grades and classroom names
        for grade in classrooms_by_grade:
            classrooms_by_grade[grade].sort()
        
        subjects = [wizard_data.subject_name] if wizard_data.subject_name else []
        grades = sorted(classrooms_by_grade.keys())
    else:
        subjects = json.loads(wizard_data.subjects) if wizard_data.subjects else []
        grades = []
        classrooms_by_grade = {}
    
    # Handle form submission
    if request.method == 'POST':
        try:
            test_id = request.form.get('test_id')
            test_scope = request.form.get('test_scope', 'grade_all')
            form_semester = (request.form.get('semester') or '').strip()
            form_grade = (request.form.get('grade') or '').strip()
            form_class_name = (request.form.get('class_name') or '').strip()
            form_subject = (request.form.get('subject') or '').strip()
            
            if test_id:
                # Update existing test
                test = Test.query.filter_by(id=test_id, teacher_id=current_user.id).first()
                if not test:
                    flash('Test not found.', 'error')
                    return redirect(url_for('main.create_tests'))
                
                test.semester = form_semester
                test.grade = form_grade
                test.class_name = form_class_name
                test.subject = form_subject
                test.competency = request.form['competency']
                test.test_name = request.form['test_name']
                test.max_points = int(request.form['max_points'])
                test.test_date = datetime.strptime(request.form['test_date'], '%Y-%m-%d').date()
                test.test_weight = float(request.form['test_weight'])
                
                flash('Test updated successfully!', 'success')
            else:
                if teacher_type == 'specialist' and test_scope == 'grade_all':
                    selected_grade = form_grade
                    class_names = []

                    try:
                        wizard_classrooms = json.loads(wizard_data.classrooms) if wizard_data.classrooms else []
                    except Exception:
                        wizard_classrooms = []

                    if selected_grade and wizard_classrooms:
                        for classroom in wizard_classrooms:
                            grade_label = str(classroom.get('grade') or '').strip()
                            classroom_name = str(classroom.get('name') or '').strip()
                            if grade_label == selected_grade and classroom_name:
                                class_names.append(classroom_name)

                    if not class_names and selected_grade:
                        class_names = classrooms_by_grade.get(selected_grade, [])

                    if not class_names:
                        try:
                            available_grades = sorted({str(c.get('grade') or '').strip() for c in (wizard_classrooms or []) if c.get('grade') is not None})
                        except Exception:
                            available_grades = []
                        current_app.logger.warning(
                            "create_tests grade_all: no class_names found", extra={
                                "teacher_id": current_user.id,
                                "selected_grade": selected_grade,
                                "available_grades": available_grades,
                            }
                        )

                    if not selected_grade or not class_names:
                        flash('Unable to create tests for the selected grade. Please re-select your Class and Semester and try again.', 'danger')
                        return redirect(url_for('main.create_tests'))

                    tests_to_create = []
                    for class_name in class_names:
                        tests_to_create.append(Test(
                            teacher_id=current_user.id,
                            semester=form_semester,
                            grade=selected_grade,
                            class_name=class_name,
                            subject=form_subject,
                            competency=request.form['competency'],
                            test_name=request.form['test_name'],
                            max_points=int(request.form['max_points']),
                            test_date=datetime.strptime(request.form['test_date'], '%Y-%m-%d').date(),
                            test_weight=float(request.form['test_weight'])
                        ))

                    db.session.add_all(tests_to_create)
                    current_app.logger.info(
                        "create_tests grade_all: creating tests", extra={
                            "teacher_id": current_user.id,
                            "selected_grade": selected_grade,
                            "class_count": len(tests_to_create),
                        }
                    )
                    flash(f'Tests created successfully! Created {len(tests_to_create)} tests.', 'success')
                else:
                    # Create new test
                    test = Test(
                        teacher_id=current_user.id,
                        semester=form_semester,
                        grade=form_grade,
                        class_name=form_class_name,
                        subject=form_subject,
                        competency=request.form['competency'],
                        test_name=request.form['test_name'],
                        max_points=int(request.form['max_points']),
                        test_date=datetime.strptime(request.form['test_date'], '%Y-%m-%d').date(),
                        test_weight=float(request.form['test_weight'])
                    )
                    
                    db.session.add(test)
                    flash('Test created successfully!', 'success')
            
            db.session.commit()
            return redirect(url_for('main.create_tests'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving test: {str(e)}', 'error')
    
    # Get existing tests for display
    tests = Test.query.filter_by(teacher_id=current_user.id).order_by(Test.test_date.desc()).all()
    
    # Get the last created test for form prepopulation
    last_test = tests[0] if tests else None
    
    return render_template('create_tests.html',
                         competencies=competencies,
                         semesters=semesters,
                         subjects=subjects,
                         grades=grades,
                         classrooms_by_grade=classrooms_by_grade,
                         teacher_type=teacher_type,
                         tests=tests,
                         last_test=last_test,
                         show_global_filters=True)

@main.route('/setup_wizard/submit', methods=['POST'])
@login_required
def submit_setup_wizard():
    from flask import request, jsonify
    from .models import School, Classroom, SetupWizardData
    import json
    
    data = request.get_json()
    teacher_type = data.get('teacher_type')
    school_name = data.get('school_name')
    num_semesters = data.get('num_semesters')
    competencies = data.get('competencies', [])
    subjects = data.get('subjects', [])
    grades = data.get('grades', [])
    weights = data.get('weights', {})
    classrooms_data = data.get('classrooms', [])
    grade_name = data.get('grade_name')
    subject_name = data.get('subject_name')  # For specialist teachers
    competencies_skipped = data.get('competencies_skipped', False)
    
    try:
        # Save or update Setup Wizard data
        wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
        if wizard_data:
            # Update existing data
            wizard_data.teacher_type = teacher_type
            wizard_data.school_name = school_name
            wizard_data.num_semesters = num_semesters
            wizard_data.competencies = json.dumps(competencies)
            wizard_data.subjects = json.dumps(subjects)
            wizard_data.grades = json.dumps(grades)
            wizard_data.weights = json.dumps(weights)
            wizard_data.classrooms = json.dumps(classrooms_data)
            wizard_data.grade_name = grade_name
            wizard_data.subject_name = subject_name
            wizard_data.competencies_skipped = competencies_skipped
        else:
            # Create new wizard data
            wizard_data = SetupWizardData(
                teacher_id=current_user.id,
                teacher_type=teacher_type,
                school_name=school_name,
                num_semesters=num_semesters,
                competencies=json.dumps(competencies),
                subjects=json.dumps(subjects),
                grades=json.dumps(grades),
                weights=json.dumps(weights),
                classrooms=json.dumps(classrooms_data),
                grade_name=grade_name,
                subject_name=subject_name,
                competencies_skipped=competencies_skipped
            )
            db.session.add(wizard_data)
        
        # Create or get the school
        school = School.query.filter_by(name=school_name, teacher_id=current_user.id).first()
        if not school:
            school = School(name=school_name, teacher_id=current_user.id)
            db.session.add(school)
            db.session.flush()  # Get the school ID
        
        # For specialist teachers, save the classrooms
        if teacher_type == 'specialist' and classrooms_data:
            for classroom_data in classrooms_data:
                classroom_name = classroom_data.get('name')
                grade = classroom_data.get('grade')
                
                # Create classroom name with grade info
                # Check if grade already contains "Grade" prefix to avoid duplication
                if grade.startswith('Grade'):
                    full_name = f"{classroom_name} ({grade})"
                else:
                    full_name = f"{classroom_name} (Grade {grade})"
                
                # Check if classroom already exists
                existing_classroom = Classroom.query.filter_by(
                    name=full_name, 
                    school_id=school.id
                ).first()
                
                if not existing_classroom:
                    classroom = Classroom(name=full_name, school_id=school.id)
                    db.session.add(classroom)
        
        # For homeroom teachers, create a single classroom
        elif teacher_type == 'homeroom':
            classroom_name = grade_name or 'My Class'
            
            # Check if classroom already exists
            existing_classroom = Classroom.query.filter_by(
                name=classroom_name, 
                school_id=school.id
            ).first()
            
            if not existing_classroom:
                classroom = Classroom(name=classroom_name, school_id=school.id)
                db.session.add(classroom)
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@main.route('/api/get_teacher_classrooms')
@login_required
def get_teacher_classrooms():
    from flask import jsonify
    from .models import School, Classroom, Student
    
    try:
        # Get all schools for the current teacher
        schools = School.query.filter_by(teacher_id=current_user.id).all()
        
        classrooms_data = []
        for school in schools:
            for classroom in school.classrooms:
                # Count students in this classroom
                student_count = Student.query.filter_by(classroom_id=classroom.id).count()
                
                classrooms_data.append({
                    'id': classroom.id,
                    'name': classroom.name,
                    'school_name': school.name,
                    'student_count': student_count
                })
        
        return jsonify({'classrooms': classrooms_data})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main.route('/export/grade_matrix.xlsx')
@login_required
def export_grade_matrix_xlsx():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from .models import Test, Student, Grade, Classroom, School, SetupWizardData

    semester = request.args.get('semester', '')
    class_name = request.args.get('class_name', '')
    subject = request.args.get('subject', '')

    def sanitize_for_filename(s: str) -> str:
        return ''.join([c if c.isalnum() else '_' for c in (s or '').strip()]).strip('_').lower()

    try:
        test_query = Test.query.filter_by(teacher_id=current_user.id)
        if semester:
            test_query = test_query.filter(Test.semester == semester)
        if class_name:
            test_query = test_query.filter(Test.class_name == class_name)
        if subject:
            test_query = test_query.filter(Test.subject == subject)

        tests = test_query.order_by(Test.test_date).all()
        if not tests:
            return jsonify({'error': 'No data to export'}), 400

        students = []
        if class_name:
            for school in School.query.filter_by(teacher_id=current_user.id).all():
                for classroom in Classroom.query.filter_by(school_id=school.id).all():
                    classroom_class_name = classroom.name.split(' (')[0] if ' (' in classroom.name else classroom.name
                    if classroom_class_name == class_name:
                        students.extend(Student.query.filter_by(classroom_id=classroom.id).all())
        else:
            for school in School.query.filter_by(teacher_id=current_user.id).all():
                for classroom in Classroom.query.filter_by(school_id=school.id).all():
                    students.extend(Student.query.filter_by(classroom_id=classroom.id).all())

        students = sorted(students, key=lambda s: (s.last_name or '', s.first_name or ''))
        if not students:
            return jsonify({'error': 'No data to export'}), 400

        test_ids = [t.id for t in tests]
        student_ids = [s.id for s in students]
        grades_query = Grade.query.filter(
            Grade.test_id.in_(test_ids),
            Grade.student_id.in_(student_ids)
        ).all()

        grades_matrix = {}
        for gr in grades_query:
            grades_matrix.setdefault(gr.student_id, {})[gr.test_id] = gr.grade

        wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
        competency_weights = {}
        if wizard_data and wizard_data.weights and wizard_data.competencies:
            try:
                weights_data = json.loads(wizard_data.weights)
                competencies_list = json.loads(wizard_data.competencies)
                for grade_key in weights_data:
                    for semester_key in weights_data[grade_key]:
                        semester_weights = weights_data[grade_key][semester_key]
                        for comp_index, weight in semester_weights.items():
                            idx = int(comp_index)
                            if idx < len(competencies_list):
                                competency_weights[competencies_list[idx]] = int(weight)
                        break
                    break
            except Exception:
                competency_weights = {}

        tests_by_comp = {}
        for t in tests:
            comp = t.competency or 'Unassigned'
            tests_by_comp.setdefault(comp, []).append(t)
        competencies = sorted(tests_by_comp.keys())

        wb = Workbook()
        ws = wb.active
        ws.title = 'Grade Matrix'

        fill_total = PatternFill('solid', fgColor='E9ECEF')
        fill_grand = PatternFill('solid', fgColor='E7F3FF')
        font_bold = Font(bold=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Row layout:
        # 1: title
        # 2: max points (tests)
        # 3: test weights (tests)
        # 4: competency weights (competency total columns)
        # 5: headers (test names)
        # 6: student-name label row
        title = 'Grade Matrix'
        if class_name or semester:
            parts = [p for p in [class_name, semester] if p]
            title = f"Grade Matrix - {' - '.join(parts)}"
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)

        max_points_row = 2
        test_weights_row = 3
        comp_weights_row = 4
        header_row = 5
        student_label_row = 6
        first_student_row = 7

        header_fill = PatternFill('solid', fgColor='D9D9D9')
        info_font = Font(color='7F7F7F')
        info_align = Alignment(horizontal='left', vertical='center')

        ws.cell(row=max_points_row, column=1, value='Test - Max Points').font = info_font
        ws.cell(row=max_points_row, column=1).alignment = info_align
        ws.cell(row=test_weights_row, column=1, value='Test Weight').font = info_font
        ws.cell(row=test_weights_row, column=1).alignment = info_align
        ws.cell(row=comp_weights_row, column=1, value='Competency Weight').font = info_font
        ws.cell(row=comp_weights_row, column=1).alignment = info_align

        col = 1
        ws.cell(row=header_row, column=col, value='Test Name').font = font_bold
        ws.cell(row=header_row, column=col).alignment = align_center
        ws.cell(row=header_row, column=col).fill = header_fill
        ws.column_dimensions[get_column_letter(col)].width = 22
        col += 1

        competency_total_cols = []
        comp_total_col_by_comp = {}
        test_cols_by_comp = {}

        for comp in competencies:
            test_cols = []
            for t in tests_by_comp[comp]:
                ws.cell(row=header_row, column=col, value=f"{t.test_name}").font = font_bold
                ws.cell(row=header_row, column=col).alignment = align_center
                ws.cell(row=header_row, column=col).fill = header_fill
                ws.cell(row=max_points_row, column=col, value=t.max_points).alignment = align_center
                ws.cell(row=max_points_row, column=col).font = info_font
                ws.cell(row=test_weights_row, column=col, value=(t.test_weight or 0)).alignment = align_center
                ws.cell(row=test_weights_row, column=col).font = info_font
                ws.column_dimensions[get_column_letter(col)].width = 12
                test_cols.append(col)
                col += 1

            # Competency total column (fraction, formatted as percent)
            ws.cell(row=header_row, column=col, value=f"Total {comp}").font = font_bold
            ws.cell(row=header_row, column=col).alignment = align_center
            ws.cell(row=header_row, column=col).fill = fill_total
            ws.column_dimensions[get_column_letter(col)].width = 14
            ws.cell(row=comp_weights_row, column=col, value=int(competency_weights.get(comp, 0))).alignment = align_center
            ws.cell(row=comp_weights_row, column=col).font = info_font
            competency_total_cols.append(col)
            comp_total_col_by_comp[comp] = col
            test_cols_by_comp[comp] = test_cols
            col += 1

        show_grand_total = len(competencies) > 1
        grand_total_col = None
        if show_grand_total:
            grand_total_col = col
            ws.cell(row=header_row, column=col, value='Grand Total').font = font_bold
            ws.cell(row=header_row, column=col).alignment = align_center
            ws.cell(row=header_row, column=col).fill = fill_grand
            ws.column_dimensions[get_column_letter(col)].width = 14
            col += 1

        last_data_col = col - 1

        ws.cell(row=student_label_row, column=1, value='Student Name').font = font_bold
        ws.cell(row=student_label_row, column=1).alignment = Alignment(horizontal='left', vertical='center')

        # Helper formulas (use standard Excel functions; keep everything as fractions for formatting)
        def comp_total_formula(student_row: int, test_cols: list[int]) -> str:
            # numerator = SUMPRODUCT((points<>"")*(points/max_points)*weights)
            # denom     = SUMPRODUCT((points<>"")*weights)
            pts = f"{get_column_letter(test_cols[0])}{student_row}:{get_column_letter(test_cols[-1])}{student_row}"
            maxs = f"{get_column_letter(test_cols[0])}{max_points_row}:{get_column_letter(test_cols[-1])}{max_points_row}"
            wts = f"{get_column_letter(test_cols[0])}{test_weights_row}:{get_column_letter(test_cols[-1])}{test_weights_row}"
            denom = f"SUMPRODUCT(({pts}<>\"\")*{wts})"
            num = f"SUMPRODUCT(({pts}<>\"\")*({pts}/{maxs})*{wts})"
            return f"=IF({denom}=0,\"\",{num}/{denom})"

        def comp_total_from_fraction_formula(row: int, fraction_cols: list[int]) -> str:
            # fraction cells already contain (points/max_points), so do NOT divide by max points again
            fr = f"{get_column_letter(fraction_cols[0])}{row}:{get_column_letter(fraction_cols[-1])}{row}"
            wts = f"{get_column_letter(fraction_cols[0])}{test_weights_row}:{get_column_letter(fraction_cols[-1])}{test_weights_row}"
            denom = f"SUMPRODUCT(({fr}<>\"\")*{wts})"
            num = f"SUMPRODUCT(({fr}<>\"\")*{fr}*{wts})"
            return f"=IF({denom}=0,\"\",{num}/{denom})"

        def grand_total_formula(student_row: int) -> str:
            comp_vals = f"{get_column_letter(competency_total_cols[0])}{student_row}:{get_column_letter(competency_total_cols[-1])}{student_row}"
            comp_wts = f"{get_column_letter(competency_total_cols[0])}{comp_weights_row}:{get_column_letter(competency_total_cols[-1])}{comp_weights_row}"
            denom = f"SUMPRODUCT(({comp_vals}<>\"\")*{comp_wts})"
            num = f"SUMPRODUCT(({comp_vals}<>\"\")*{comp_vals}*{comp_wts})"
            return f"=IF({denom}=0,\"\",{num}/{denom})"

        # Student rows
        for i, s in enumerate(students):
            r = first_student_row + i
            ws.cell(row=r, column=1, value=f"{s.first_name} {s.last_name}")
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')

            for comp in competencies:
                # test points (editable)
                for t, test_col in zip(tests_by_comp[comp], test_cols_by_comp[comp]):
                    val = grades_matrix.get(s.id, {}).get(t.id)
                    if val is not None:
                        ws.cell(row=r, column=test_col, value=float(val))
                    else:
                        ws.cell(row=r, column=test_col, value=0)
                    ws.cell(row=r, column=test_col).alignment = align_center

                # competency total
                total_col = comp_total_col_by_comp[comp]
                f = comp_total_formula(r, test_cols_by_comp[comp])
                cell = ws.cell(row=r, column=total_col, value=f)
                cell.number_format = '0.0%'
                cell.font = font_bold
                cell.fill = fill_total
                cell.alignment = align_center

            if show_grand_total and grand_total_col:
                gcell = ws.cell(row=r, column=grand_total_col, value=grand_total_formula(r))
                gcell.number_format = '0.0%'
                gcell.font = font_bold
                gcell.fill = fill_grand
                gcell.alignment = align_center

        # Class average row
        avg_row = first_student_row + len(students)
        ws.cell(row=avg_row, column=1, value='Class Average').font = font_bold
        ws.cell(row=avg_row, column=1).alignment = Alignment(horizontal='left', vertical='center')

        for comp in competencies:
            # per-test avg as fraction
            for test_col in test_cols_by_comp[comp]:
                col_letter = get_column_letter(test_col)
                start = first_student_row
                end = first_student_row + len(students) - 1
                max_ref = f"{col_letter}{max_points_row}"
                f = f"=IF(COUNT({col_letter}{start}:{col_letter}{end})=0,\"\",AVERAGE({col_letter}{start}:{col_letter}{end})/{max_ref})"
                cell = ws.cell(row=avg_row, column=test_col, value=f)
                cell.number_format = '0.0%'
                cell.alignment = align_center

            total_col = comp_total_col_by_comp[comp]
            f = comp_total_from_fraction_formula(avg_row, test_cols_by_comp[comp])
            cell = ws.cell(row=avg_row, column=total_col, value=f)
            cell.number_format = '0.0%'
            cell.font = font_bold
            cell.fill = fill_total
            cell.alignment = align_center

        if show_grand_total and grand_total_col:
            cell = ws.cell(row=avg_row, column=grand_total_col, value=grand_total_formula(avg_row))
            cell.number_format = '0.0%'
            cell.font = font_bold
            cell.fill = fill_grand
            cell.alignment = align_center

        # Cosmetics: center alignment for meta/header rows
        for r in [header_row, max_points_row, test_weights_row, comp_weights_row]:
            for c in range(1, last_data_col + 1):
                cell = ws.cell(row=r, column=c)
                if r == header_row:
                    cell.font = font_bold
                    cell.alignment = align_center
                else:
                    if c != 1:
                        cell.alignment = align_center

        ws.freeze_panes = ws['B7']

        filename_parts = ['grade_matrix']
        if class_name:
            filename_parts.append(sanitize_for_filename(class_name))
        if semester:
            filename_parts.append(sanitize_for_filename(semester))
        filename = '_'.join([p for p in filename_parts if p]) + '.xlsx'

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        current_app.logger.exception('Error exporting grade matrix to xlsx')
        return jsonify({'error': str(e)}), 500

@main.route('/api/get_tests_for_context')
@login_required
def get_tests_for_context():
    """Return tests for the current teacher filtered by semester and optional class/subject.
    Query params:
      - semester (required)
      - class_name (optional, for specialist)
      - subject (optional, for homeroom)
    """
    from .models import Test, SetupWizardData
    try:
        semester = request.args.get('semester', '').strip()
        class_name = request.args.get('class_name', '').strip()
        subject = request.args.get('subject', '').strip()

        if not semester:
            return jsonify({'error': 'semester is required'}), 400

        wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
        teacher_type = wizard_data.teacher_type if wizard_data and wizard_data.teacher_type else 'homeroom'

        q = Test.query.filter_by(teacher_id=current_user.id, semester=semester)
        if teacher_type == 'specialist':
            if class_name:
                q = q.filter(Test.class_name == class_name)
        else:
            if subject:
                q = q.filter(Test.subject == subject)

        tests = q.order_by(Test.test_date.desc()).all()
        return jsonify({'tests': [
            {
                'id': t.id,
                'test_name': t.test_name,
                'test_date': t.test_date.strftime('%Y-%m-%d'),
                'max_points': t.max_points,
                'competency': t.competency,
                'test_weight': t.test_weight
            } for t in tests
        ]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main.route('/api/bell_grade_scenarios', methods=['POST'])
@login_required
def bell_grade_scenarios():
    """Compute bell grading scenarios for a given test.
    Body JSON:
      - test_id (int, required)
      - adjust_avg (bool)
      - target_avg (float|null)
      - allow_over_100 (bool)
      - boost_low (bool)
      - lowest_score (float|null)
    Returns per-student original percentage and selected scenario percentages.
    """
    from .models import Test, Grade, Student, Classroom, School
    try:
        data = request.get_json(force=True) or {}
        test_id = data.get('test_id')
        if not test_id:
            return jsonify({'error': 'test_id is required'}), 400

        adjust_avg = bool(data.get('adjust_avg'))
        target_avg = data.get('target_avg')
        allow_over_100 = bool(data.get('allow_over_100'))
        boost_low = bool(data.get('boost_low'))
        lowest_score = data.get('lowest_score')

        test = Test.query.filter_by(id=test_id, teacher_id=current_user.id).first()
        if not test:
            return jsonify({'error': 'Test not found'}), 404

        # Fetch (grade, student) tuples for this test, ensuring students belong to the current teacher
        class_name = (data.get('class_name') or '').strip()
        q = (
            db.session.query(Grade, Student)
            .join(Student, Grade.student_id == Student.id)
            .join(Classroom, Student.classroom_id == Classroom.id)
            .join(School, Classroom.school_id == School.id)
            .filter(Grade.test_id == test.id, School.teacher_id == current_user.id)
        )
        if class_name:
            # Match classrooms that start with the class name (e.g., '101' matches '101 (Grade 1)')
            q = q.filter(Classroom.name.like(f"{class_name}%"))
        grade_rows = q.all()

        # Build original percentages and compute class average
        students_map = {}
        percentages = []
        for g, stu in grade_rows:
            name = f"{stu.first_name} {stu.last_name}"
            if g.absent or g.grade is None:
                pct = None
            else:
                pct = (g.grade / test.max_points) * 100.0
                percentages.append(pct)
            students_map[g.student_id] = {
                'name': name,
                'original': pct
            }

        # Compute class average of available grades
        original_class_avg = sum(percentages) / len(percentages) if percentages else None

        # Debug logging
        print(f"DEBUG BELL_SCENARIOS: test_id={test_id}, adjust_avg={adjust_avg}, target_avg={target_avg}")
        print(f"DEBUG BELL_SCENARIOS: grade_rows count={len(grade_rows)}, percentages count={len(percentages)}")
        print(f"DEBUG BELL_SCENARIOS: original_class_avg={original_class_avg}")

        # Prepare helpers
        def cap100(val: float) -> float:
            return min(val, 100.0)

        # Compute scenarios based on selections
        if adjust_avg and (target_avg is None or original_class_avg is None):
            print(f"DEBUG BELL_SCENARIOS: FAILING - adjust_avg={adjust_avg}, target_avg={target_avg}, original_class_avg={original_class_avg}")
            return jsonify({'error': 'Target average invalid or no graded data available for adjustment'}), 400

        if isinstance(target_avg, (int, float)):
            target_avg = float(target_avg)
        if isinstance(lowest_score, (int, float)):
            lowest_score = float(lowest_score)

        linear_diff = None
        ratio = None
        if adjust_avg and original_class_avg is not None:
            linear_diff = (target_avg or 0.0) - original_class_avg
            ratio = (target_avg or 0.0) / original_class_avg if original_class_avg != 0 else None

        # Build response list maintaining name order by student last/first
        # Sort keys by name for consistent display
        students_list = sorted(students_map.values(), key=lambda s: s['name'].split(' ')[-1] + ' ' + s['name'].split(' ')[0])
        response_students = []
        for s in students_list:
            original = s['original']
            out = {
                'name': s['name'],
                'original': original if original is not None else None,
                'linear': None,
                'percentage': None,
                'sqrt': None,
            }

            if original is not None:
                # Equal Adjustment: flat adjustment to all scores
                if adjust_avg and linear_diff is not None:
                    val = original + linear_diff
                    if not allow_over_100:
                        val = cap100(val)
                    out['linear'] = val

                # Lower Boost: square root boost for lower scores
                if boost_low and original_class_avg is not None:
                    adjust = math.sqrt(max(original_class_avg, 0.0)) * 10.0 - target_avg
                    val = math.sqrt(max(original, 0.0)) * 10.0 - adjust
                    if lowest_score is not None:
                        val = max(val, lowest_score)
                    # Ensure scores never go down - use max of (original + 2) or calculated value
                    val = max(original + 2.0, val)
                    out['sqrt'] = val

                # Typical Distribution: compress and lift
                if adjust_avg and original_class_avg is not None:
                    ratio = 0.85
                    uplift = target_avg - (original_class_avg * ratio)
                    val = (original * ratio) + uplift
                    if not allow_over_100:
                        val = cap100(val)
                    out['percentage'] = val

            response_students.append(out)

        return jsonify({
            'test': {
                'id': test.id,
                'name': test.test_name,
                'max_points': test.max_points,
                'original_class_avg': original_class_avg
            },
            'students': response_students
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main.route('/api/apply_bell_selection', methods=['POST'])
@login_required
def apply_bell_selection():
    """Apply the selected bell grading scenario to persist grades.
    Body JSON expects the same options as scenarios plus 'scenario' key in ['original','linear','percentage','sqrt'].
    This updates Grade.grade (points) for each student and marks the Test as modified with details.
    """
    from .models import Test, Grade, Student, Classroom, School
    try:
        data = request.get_json(force=True) or {}
        test_id = data.get('test_id')
        scenario = data.get('scenario')
        if not test_id or scenario not in ['original','linear','percentage','sqrt']:
            return jsonify({'error': 'Invalid request'}), 400

        adjust_avg = bool(data.get('adjust_avg'))
        target_avg = data.get('target_avg')
        allow_over_100 = bool(data.get('allow_over_100'))
        boost_low = bool(data.get('boost_low'))
        lowest_score = data.get('lowest_score')
        class_name = (data.get('class_name') or '').strip()

        test = Test.query.filter_by(id=test_id, teacher_id=current_user.id).first()
        if not test:
            return jsonify({'error': 'Test not found'}), 404

        # Query relevant students/grades
        q = (
            db.session.query(Grade, Student)
            .join(Student, Grade.student_id == Student.id)
            .join(Classroom, Student.classroom_id == Classroom.id)
            .join(School, Classroom.school_id == School.id)
            .filter(Grade.test_id == test.id, School.teacher_id == current_user.id)
        )
        if class_name:
            # Match classroom names that start with the class_name (e.g., "101" matches "101 (Grade 1)")
            q = q.filter(Classroom.name.like(f"{class_name}%"))
        rows = q.all()

        # Build original percentages and class average
        percentages = []
        for g, _stu in rows:
            if not g.absent and g.grade is not None:
                percentages.append((g.grade / test.max_points) * 100.0)
        original_class_avg = sum(percentages) / len(percentages) if percentages else None

        if isinstance(target_avg, (int, float)):
            target_avg = float(target_avg)
        if isinstance(lowest_score, (int, float)):
            lowest_score = float(lowest_score)

        # Compute adjustment params
        linear_diff = None
        ratio = None
        if adjust_avg and original_class_avg is not None:
            linear_diff = (target_avg or 0.0) - original_class_avg
            ratio = (target_avg or 0.0) / original_class_avg if original_class_avg != 0 else None

        def cap100(val: float) -> float:
            return min(val, 100.0)

        # Apply scenario per student
        updated = 0
        for g, _stu in rows:
            if g.absent or g.grade is None:
                continue
            orig_pct = (g.grade / test.max_points) * 100.0
            new_pct = None
            if scenario == 'original':
                new_pct = orig_pct
            elif scenario == 'linear' and linear_diff is not None:
                new_pct = orig_pct + linear_diff
                if not allow_over_100:
                    new_pct = cap100(new_pct)
            elif scenario == 'percentage' and ratio is not None:
                new_pct = orig_pct * ratio
                if not allow_over_100:
                    new_pct = cap100(new_pct)
            elif scenario == 'sqrt' and boost_low and original_class_avg is not None:
                adjust = math.sqrt(max(original_class_avg, 0.0)) * 10.0 - target_avg
                new_pct = math.sqrt(max(orig_pct, 0.0)) * 10.0 - adjust
                if lowest_score is not None:
                    new_pct = max(new_pct, lowest_score)
                # Ensure scores never go down - use max of (original + 2) or calculated value
                new_pct = max(orig_pct + 2.0, new_pct)

            if new_pct is None:
                continue
            # Convert percent back to points
            new_points = (new_pct / 100.0) * test.max_points
            g.grade = round(new_points, 2)
            
            # Set modification tracking fields
            scenario_names = {
                'linear': 'Equal Adjustment',
                'sqrt': 'Lower Boost',
                'percentage': 'Typical Distribution',
                'original': 'Original'
            }
            g.modification_type = 'bell_grade'
            g.modification_notes = f"Bell Grade - {scenario_names.get(scenario, scenario)}, Target Avg: {target_avg}%"
            g.modified_at = datetime.utcnow()
            
            updated += 1

        # Persist changes
        details = (
            f"Type: {'Linear scaling' if scenario=='linear' else ('% scaling' if scenario=='percentage' else ('Square root' if scenario=='sqrt' else 'Original'))}; "
            f"Original average: {original_class_avg if original_class_avg is not None else 'N/A'}; "
            f"New average: {target_avg if adjust_avg and target_avg is not None else 'N/A'}; "
            f"Scores > 100% allowed: {'True' if allow_over_100 else 'False'}; "
            f"Lowest score allowed: {lowest_score if lowest_score is not None else 'N/A'}"
        )
        test.scores_modified = True
        test.scores_modified_details = details
        db.session.commit()

        return jsonify({'updated': updated, 'test_id': test.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@main.route('/api/save_students', methods=['POST'])
@login_required
def save_students():
    from flask import request, jsonify
    from .models import Student, Classroom, School
    
    data = request.get_json()
    classroom_id = data.get('classroom_id')
    students_data = data.get('students', [])
    
    try:
        # Verify classroom belongs to current teacher
        classroom = Classroom.query.join(School).filter(
            Classroom.id == classroom_id,
            School.teacher_id == current_user.id
        ).first()
        
        if not classroom:
            return jsonify({'success': False, 'error': 'Classroom not found'}), 404
        
        # Get existing students in this classroom
        existing_students = Student.query.filter_by(classroom_id=classroom_id).all()
        existing_students_map = {(s.first_name.lower(), s.last_name.lower()): s for s in existing_students}
        
        # Track which students from the incoming data already exist
        incoming_student_keys = set()
        new_students_added = 0
        
        # Add or update students from incoming data
        for student_data in students_data:
            first_name = student_data.get('firstName', '').strip()
            last_name = student_data.get('lastName', '').strip()
            student_key = (first_name.lower(), last_name.lower())
            incoming_student_keys.add(student_key)
            
            # Only add if student doesn't already exist
            if student_key not in existing_students_map:
                student = Student(
                    first_name=first_name,
                    last_name=last_name,
                    classroom_id=classroom_id
                )
                db.session.add(student)
                new_students_added += 1
        
        db.session.commit()
        return jsonify({'success': True, 'count': new_students_added, 'total': len(students_data)})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@main.route('/api/get_classroom_students/<int:classroom_id>')
@login_required
def get_classroom_students(classroom_id):
    """Get all students in a specific classroom"""
    from flask import jsonify
    from .models import Student, Classroom, School
    
    try:
        # Verify classroom belongs to current teacher
        classroom = Classroom.query.join(School).filter(
            Classroom.id == classroom_id,
            School.teacher_id == current_user.id
        ).first()
        
        if not classroom:
            return jsonify({'success': False, 'error': 'Classroom not found'}), 404
        
        # Get all students in this classroom
        students = Student.query.filter_by(classroom_id=classroom_id).order_by(
            Student.last_name.asc(), Student.first_name.asc()
        ).all()
        
        students_data = []
        for student in students:
            students_data.append({
                'id': student.id,
                'firstName': student.first_name,
                'lastName': student.last_name,
                'fullName': f"{student.first_name} {student.last_name}"
            })
        
        return jsonify({
            'success': True,
            'students': students_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@main.route('/api/get_setup_wizard_data')
@login_required
def get_setup_wizard_data():
    from flask import jsonify
    from .models import SetupWizardData
    import json
    
    try:
        wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
        
        if wizard_data:
            # Parse JSON fields back to Python objects
            competencies = json.loads(wizard_data.competencies) if wizard_data.competencies else []
            subjects = json.loads(wizard_data.subjects) if wizard_data.subjects else []
            grades = json.loads(wizard_data.grades) if wizard_data.grades else []
            weights = json.loads(wizard_data.weights) if wizard_data.weights else {}
            classrooms = json.loads(wizard_data.classrooms) if wizard_data.classrooms else []
            
            return jsonify({
                'success': True,
                'has_previous_data': True,
                'data': {
                    'teacher_type': wizard_data.teacher_type,
                    'school_name': wizard_data.school_name,
                    'num_semesters': wizard_data.num_semesters,
                    'competencies': competencies,
                    'subjects': subjects,
                    'grades': grades,
                    'weights': weights,
                    'classrooms': classrooms,
                    'grade_name': wizard_data.grade_name,
                    'subject_name': wizard_data.subject_name
                }
            })
        else:
            return jsonify({
                'success': True,
                'has_previous_data': False,
                'data': None
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@main.route('/api/get_student_data/<int:student_id>')
@login_required
def get_student_data(student_id):
    """Get detailed student data including grades and photo"""
    from flask import jsonify
    from .models import Student, Grade, Test, Classroom, School
    import random
    
    try:
        # Verify student belongs to current teacher
        student = Student.query.join(Classroom).join(School).filter(
            Student.id == student_id,
            School.teacher_id == current_user.id
        ).first()
        
        if not student:
            return jsonify({'success': False, 'error': 'Student not found'}), 404
        
        # Get all grades for this student
        grades = Grade.query.join(Test).filter(
            Grade.student_id == student_id,
            Test.teacher_id == current_user.id
        ).order_by(Test.test_date.asc()).all()
        
        # Process grades data
        grades_data = []
        low_grades_count = 0
        
        for grade in grades:
            if grade.absent:
                grade_percentage = None
                status = 'Absent'
            elif grade.grade is not None:
                grade_percentage = round((grade.grade / grade.test.max_points) * 100, 1)
                status = 'Completed'
                if grade_percentage < 80:
                    low_grades_count += 1
            else:
                grade_percentage = None
                status = 'Not Graded'
            
            grades_data.append({
                'test_name': grade.test.test_name,
                'subject': grade.test.subject,
                'competency': grade.test.competency,
                'test_date': grade.test.test_date.strftime('%Y-%m-%d'),
                'semester': grade.test.semester,
                'grade': grade.grade,
                'max_points': grade.test.max_points,
                'percentage': grade_percentage,
                'status': status,
                'absent': grade.absent
            })
        
        # Generate random avatar (placeholder for photo functionality)
        avatar_options = [
            'https://api.dicebear.com/7.x/avataaars/svg?seed=' + student.first_name + student.last_name,
            'https://api.dicebear.com/7.x/personas/svg?seed=' + student.first_name + student.last_name,
            'https://api.dicebear.com/7.x/fun-emoji/svg?seed=' + student.first_name + student.last_name
        ]
        avatar_url = random.choice(avatar_options)
        
        # Sample teacher notes (placeholder)
        sample_notes = [
            f"{student.first_name} shows excellent participation in class discussions.",
            f"Strong analytical skills demonstrated in recent assignments.",
            f"Would benefit from additional practice with problem-solving techniques.",
            f"Consistently submits high-quality work on time.",
            f"Great improvement shown in recent assessments."
        ]
        
        return jsonify({
            'success': True,
            'student': {
                'id': student.id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'full_name': f"{student.first_name} {student.last_name}",
                'classroom': student.classroom.name,
                'avatar_url': avatar_url
            },
            'grades': grades_data,
            'stats': {
                'total_tests': len(grades_data),
                'low_grades_count': low_grades_count,
                'has_low_grades': low_grades_count > 0
            },
            'teacher_notes': sample_notes[:3]  # Show first 3 sample notes
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@main.route('/api/get_teacher_type')
@login_required
def get_teacher_type():
    from flask import jsonify
    from .models import Teacher, SetupWizardData
    
    try:
        # First, try to get teacher type from Setup Wizard data
        wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
        if wizard_data and wizard_data.teacher_type:
            return jsonify({
                'success': True,
                'teacher_type': wizard_data.teacher_type
            })
        
        # Fallback: if no setup wizard data, use heuristic based on classrooms
        teacher = Teacher.query.get(current_user.id)
        if teacher:
            total_classrooms = 0
            for school in teacher.schools:
                total_classrooms += len(school.classrooms)
            
            # Simple heuristic: if more than 1 classroom, assume Specialist
            teacher_type = 'specialist' if total_classrooms > 1 else 'homeroom'
            
            return jsonify({
                'success': True,
                'teacher_type': teacher_type
            })
        else:
            return jsonify({'success': False, 'error': 'Teacher not found'}), 404
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@login_manager.user_loader
def load_user(user_id):
    return Teacher.query.get(int(user_id))


@main.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    form = RegistrationForm()
    if form.validate_on_submit():
        existing = Teacher.query.filter_by(email=form.email.data).first()
        if existing:
            flash('Email already registered.', 'danger')
            return redirect(url_for('main.register'))
        teacher = Teacher(
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            email=form.email.data,
            password_hash=generate_password_hash(form.password.data)
        )
        db.session.add(teacher)
        db.session.commit()
        flash('Account created! Please log in.', 'success')
        return redirect(url_for('main.login'))
    return render_template('register.html', form=form)

@main.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    form = LoginForm()
    if form.validate_on_submit():
        teacher = Teacher.query.filter_by(email=form.email.data).first()
        if teacher and check_password_hash(teacher.password_hash, form.password.data):
            login_user(teacher)
            return redirect(url_for('main.dashboard'))
        flash('Invalid email or password.', 'danger')
    return render_template('login.html', form=form)

@main.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('main.index'))

@main.route('/flush_database', methods=['POST'])
@login_required
def flush_database():
    """Flush all data for the current logged-in user (for testing purposes)"""
    try:
        from .models import School, Classroom, Student, SetupWizardData, Test, Grade
        
        # Get current user ID
        user_id = current_user.id
        
        # Delete all students for this user's classrooms
        user_schools = School.query.filter_by(teacher_id=user_id).all()
        for school in user_schools:
            classrooms = Classroom.query.filter_by(school_id=school.id).all()
            for classroom in classrooms:
                # Delete all students in this classroom
                Student.query.filter_by(classroom_id=classroom.id).delete()
            # Delete all classrooms in this school
            Classroom.query.filter_by(school_id=school.id).delete()
        
        # Delete all schools for this user
        School.query.filter_by(teacher_id=user_id).delete()
        
        # Delete Setup Wizard data for this user
        SetupWizardData.query.filter_by(teacher_id=user_id).delete()
        
        # Delete all grades for tests created by this user
        user_tests = Test.query.filter_by(teacher_id=user_id).all()
        for test in user_tests:
            Grade.query.filter_by(test_id=test.id).delete()
        
        # Delete all tests for this user
        Test.query.filter_by(teacher_id=user_id).delete()
        
        # Commit all deletions
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Database flushed successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@main.route('/classroom')
@login_required
def classroom():
    from .models import SetupWizardData, Classroom, School
    
    # Get Setup Wizard data for the current user
    wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
    
    if not wizard_data:
        flash('Please complete the Setup Wizard first.', 'warning')
        return redirect(url_for('main.setup_wizard'))
    
    # Parse wizard data
    competencies = json.loads(wizard_data.competencies) if wizard_data.competencies else []
    semesters = [f"Semester {i}" for i in range(1, wizard_data.num_semesters + 1)] if wizard_data.num_semesters else []
    teacher_type = wizard_data.teacher_type if wizard_data.teacher_type else 'homeroom'
    
    # Organize data based on teacher type
    if teacher_type == 'specialist':
        # Get all classrooms for specialist teachers
        classrooms = Classroom.query.join(School).filter(School.teacher_id == current_user.id).all()
        classrooms_by_grade = {}
        
        # Parse the classrooms data from setup wizard to get grade information
        setup_classrooms = json.loads(wizard_data.classrooms) if wizard_data.classrooms else []
        
        for classroom in classrooms:
            classroom_name = classroom.name
            
            # Find the grade from setup wizard data
            grade = None
            for setup_classroom in setup_classrooms:
                if setup_classroom.get('name') == classroom_name:
                    grade = setup_classroom.get('grade')
                    break
            
            if grade:
                if grade not in classrooms_by_grade:
                    classrooms_by_grade[grade] = []
                classrooms_by_grade[grade].append(classroom_name)
        
        # Sort grades and classroom names
        for grade in classrooms_by_grade:
            classrooms_by_grade[grade].sort()
        
        subjects = json.loads(wizard_data.subjects) if wizard_data.subjects else []
        grades = sorted(classrooms_by_grade.keys())
    else:
        subjects = [wizard_data.subject_name] if wizard_data.subject_name else []
        grades = []
        classrooms_by_grade = {}
    
    return render_template('classroom.html',
                         competencies=competencies,
                         semesters=semesters,
                         subjects=subjects,
                         grades=grades,
                         classrooms_by_grade=classrooms_by_grade,
                         teacher_type=teacher_type,
                         show_global_filters=True)

@main.route('/api/save_classroom_layout', methods=['POST'])
@login_required
def save_classroom_layout():
    try:
        data = request.get_json()
        classroom_id = data.get('classroom_id')
        layout_data = data.get('layout_data')
        
        if not classroom_id or not layout_data:
            return jsonify({'success': False, 'error': 'Missing classroom_id or layout_data'}), 400
        
        # Check if layout already exists for this teacher and classroom
        existing_layout = ClassroomLayout.query.filter_by(
            teacher_id=current_user.id,
            classroom_id=classroom_id
        ).first()
        
        if existing_layout:
            # Update existing layout
            existing_layout.layout_data = json.dumps(layout_data)
            existing_layout.updated_at = datetime.utcnow()
        else:
            # Create new layout
            new_layout = ClassroomLayout(
                teacher_id=current_user.id,
                classroom_id=classroom_id,
                layout_data=json.dumps(layout_data)
            )
            db.session.add(new_layout)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Classroom layout saved successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@main.route('/api/get_classroom_layout/<int:classroom_id>')
@login_required
def get_classroom_layout(classroom_id):
    try:
        layout = ClassroomLayout.query.filter_by(
            teacher_id=current_user.id,
            classroom_id=classroom_id
        ).first()
        
        if layout:
            return jsonify({
                'success': True,
                'layout_data': json.loads(layout.layout_data)
            })
        else:
            return jsonify({
                'success': True,
                'layout_data': None
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@main.route('/preferences')
@login_required
def preferences():
    """Preferences page for application settings"""
    return render_template('preferences.html')

@main.route('/dashboard')
@login_required
def dashboard():
    from .models import School, Classroom, Test, Grade, SetupWizardData, Student
    
    # Get all schools for this teacher
    schools = School.query.filter_by(teacher_id=current_user.id).all()
    
    # Determine if Setup Wizard is completed
    setup_completed = len(schools) > 0
    
    # Initialize dashboard data
    dashboard_data = {
        'setup_completed': setup_completed,
        'schools': schools
    }
    
    if setup_completed:
        # Get detailed classroom information first
        all_classrooms = []
        classrooms_by_school = {}
        
        for school in schools:
            school_classrooms = Classroom.query.filter_by(school_id=school.id).all()
            classrooms_by_school[school.name] = school_classrooms
            all_classrooms.extend(school_classrooms)
        
        # Determine teacher type based on classroom count
        total_classrooms = len(all_classrooms)
        teacher_type = 'specialist' if total_classrooms > 1 else 'homeroom'
        
        # For specialist teachers, organize classrooms by grade
        if teacher_type == 'specialist':
            # Group classrooms by grade - extract grade and classroom name from stored format
            classrooms_by_grade = {}
            for classroom in all_classrooms:
                # Extract classroom name and grade from formats like "102 (Grade 1)" or just "102"
                if '(' in classroom.name and ')' in classroom.name:
                    # Format: "102 (Grade 1)" -> classroom_name="102", grade="Grade 1"
                    parts = classroom.name.split(' (')
                    classroom_name = parts[0]
                    grade = parts[1].rstrip(')')
                else:
                    # Format: "102" -> extract grade from classroom name
                    classroom_name = classroom.name
                    grade = extract_grade_from_classroom_name(classroom.name)
                
                if grade not in classrooms_by_grade:
                    classrooms_by_grade[grade] = []
                classrooms_by_grade[grade].append(classroom_name)
            
            # Sort classroom names within each grade A-Z
            for grade in classrooms_by_grade:
                classrooms_by_grade[grade].sort()
            
            dashboard_data['classrooms_by_grade'] = classrooms_by_grade
            
            # Get specialist subject from Setup Wizard data
            wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
            if wizard_data and wizard_data.subject_name:
                dashboard_data['specialist_subject'] = wizard_data.subject_name
            else:
                dashboard_data['specialist_subject'] = None
        
        # For homeroom teachers, get actual subjects from Setup Wizard data
        elif teacher_type == 'homeroom' and all_classrooms:
            # Get Setup Wizard data to retrieve actual subjects and grade name
            wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
            if wizard_data:
                # Use actual subjects from Setup Wizard
                dashboard_data['subjects'] = json.loads(wizard_data.subjects) if wizard_data.subjects else []
                # Use grade name from Setup Wizard (e.g., "Grade 3")
                dashboard_data['grade'] = wizard_data.grade_name if wizard_data.grade_name else all_classrooms[0].name
            else:
                # Fallback to default subjects if no Setup Wizard data
                dashboard_data['subjects'] = ['Mathematics', 'English', 'Science', 'Social Studies']
                dashboard_data['grade'] = all_classrooms[0].name
        
        # Check for missing students and tests notifications
        notifications = []
        
        # Check if competencies were skipped in setup wizard
        wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
        if wizard_data and wizard_data.competencies_skipped:
            notifications.append({
                'type': 'competencies',
                'message': _('You have not filled out your competency information yet. You will need to do this for the grade calculations to work correctly'),
                'button_text': _('Complete Setup'),
                'button_url': url_for('main.setup_wizard')
            })
        
        # Check for classes without students
        classes_without_students = []
        for classroom in all_classrooms:
            student_count = Student.query.filter_by(classroom_id=classroom.id).count()
            if student_count == 0:
                # Extract clean classroom name for display
                if '(' in classroom.name and ')' in classroom.name:
                    clean_name = classroom.name.split(' (')[0]
                else:
                    clean_name = classroom.name
                classes_without_students.append(clean_name)
        
        if classes_without_students:
            class_names = ', '.join(classes_without_students)
            notifications.append({
                'type': 'students',
                'message': _('You have not inputted students for class %(class_names)s', class_names=class_names),
                'button_text': _("Let's go!"),
                'button_url': url_for('main.student_wizard')
            })
        
        # Check for classes without tests (missing competency coverage)
        if wizard_data:
            competencies = json.loads(wizard_data.competencies) if wizard_data.competencies else []
            classes_without_tests = []
            
            for classroom in all_classrooms:
                # Extract clean classroom name for display
                if '(' in classroom.name and ')' in classroom.name:
                    clean_name = classroom.name.split(' (')[0]
                else:
                    clean_name = classroom.name
                
                # Check if this class has tests for all competencies
                if teacher_type == 'homeroom':
                    # For homeroom, check tests by subject
                    subjects = json.loads(wizard_data.subjects) if wizard_data.subjects else []
                    has_tests = False
                    for subject in subjects:
                        test_count = Test.query.filter_by(
                            teacher_id=current_user.id,
                            subject=subject
                        ).count()
                        if test_count > 0:
                            has_tests = True
                            break
                    if not has_tests:
                        classes_without_tests.append(clean_name)
                        
                elif teacher_type == 'specialist':
                    # For specialist, check tests by class
                    test_count = Test.query.filter_by(
                        teacher_id=current_user.id,
                        class_name=clean_name
                    ).count()
                    if test_count == 0:
                        classes_without_tests.append(clean_name)
            
            if classes_without_tests:
                class_names = ', '.join(classes_without_tests)
                notifications.append({
                    'type': 'tests',
                    'message': _('You have not created any tests for class %(class_names)s', class_names=class_names),
                    'button_text': _("Let's go!"),
                    'button_url': url_for('main.create_tests')
                })
        
        # Check for overdue tests (older than 1 week without grades)
        from datetime import datetime, timedelta
        one_week_ago = datetime.now() - timedelta(days=7)
        
        # Find tests older than 1 week that have no grades entered
        overdue_tests = Test.query.filter(
            Test.teacher_id == current_user.id,
            Test.test_date < one_week_ago
        ).outerjoin(Grade).filter(Grade.id == None).order_by(Test.test_date.asc()).all()
        
        if overdue_tests:
            overdue_count = len(overdue_tests)
            oldest_test = overdue_tests[0]  # First in ascending order is oldest
            
            notifications.append({
                'type': 'overdue_grading',
                'message': _('You have %(count)d tests that occurred over a week ago which need marking and grades inputted', count=overdue_count),
                'button_text': _("Grade Now"),
                'button_url': url_for('main.input_grades', test_id=oldest_test.id)
            })
        
        # Get makeup tests data (students who were absent for tests)
        makeup_tests = []
        if setup_completed:
            # Query for grades marked as absent, joined with student and test data
            absent_grades = db.session.query(Grade, Student, Test).join(
                Student, Grade.student_id == Student.id
            ).join(
                Test, Grade.test_id == Test.id
            ).filter(
                Test.teacher_id == current_user.id,
                Grade.absent == True
            ).order_by(Test.test_date, Student.last_name).all()  # Sort by date first, then student name
            
            for grade, student, test in absent_grades:
                makeup_tests.append({
                    'class_name': test.class_name,
                    'student_name': f"{student.first_name} {student.last_name}",
                    'test_name': test.test_name,
                    'test_date': test.test_date
                })
        
        # Get ungraded tests (tests with no grades entered)
        ungraded_tests = []
        if setup_completed:
            all_tests = Test.query.filter_by(teacher_id=current_user.id).all()
            for test in all_tests:
                # Check if test has any grades
                has_grades = Grade.query.filter_by(test_id=test.id).first() is not None
                if not has_grades:
                    ungraded_tests.append({
                        'class_name': test.class_name or 'N/A',
                        'test_name': test.test_name,
                        'test_date': test.test_date,
                        'status': 'Not graded'
                    })
            # Sort by date (oldest first)
            ungraded_tests.sort(key=lambda x: x['test_date'])
        
        # Get upcoming tests (tests with future dates or today's date)
        upcoming_tests = []
        if setup_completed:
            today = date.today()
            future_tests = Test.query.filter(
                Test.teacher_id == current_user.id,
                Test.test_date >= today
            ).order_by(Test.test_date).all()
            
            for test in future_tests:
                upcoming_tests.append({
                    'class_name': test.class_name or 'N/A',
                    'test_name': test.test_name,
                    'test_date': test.test_date,
                    'status': 'Upcoming'
                })
        
        dashboard_data.update({
            'teacher_type': teacher_type,
            'classrooms_by_school': classrooms_by_school,
            'all_classrooms': all_classrooms,
            'notifications': notifications,
            'makeup_tests': makeup_tests,
            'ungraded_tests': ungraded_tests,
            'upcoming_tests': upcoming_tests
        })
    
    return render_template('dashboard.html', **dashboard_data)

@main.route('/dashboard/select_school/<int:school_id>')
@login_required
def select_school(school_id):
    from .models import School, Classroom
    school = School.query.filter_by(id=school_id, teacher_id=current_user.id).first_or_404()
    classrooms = Classroom.query.filter_by(school_id=school.id).all()
    return render_template('school_dashboard.html', school=school, classrooms=classrooms)

@main.route('/schools', methods=['GET', 'POST'])
@login_required
def manage_schools():
    from .models import School
    from .forms import SchoolForm
    schools = School.query.filter_by(teacher_id=current_user.id).all()
    form = SchoolForm()
    if form.validate_on_submit():
        if len(schools) >= 5:
            flash('Maximum of 5 schools allowed.', 'danger')
        elif any(s.name == form.name.data for s in schools):
            flash('School name must be unique.', 'danger')
        else:
            new_school = School(name=form.name.data, teacher_id=current_user.id)
            db.session.add(new_school)
            db.session.commit()
            flash('School added!', 'success')
            return redirect(url_for('main.manage_schools'))
    return render_template('schools.html', form=form, schools=schools)

@main.route('/schools/edit/<int:school_id>', methods=['GET', 'POST'])
@login_required
def edit_school(school_id):
    from .models import School
    from .forms import EditSchoolForm
    school = School.query.filter_by(id=school_id, teacher_id=current_user.id).first_or_404()
    form = EditSchoolForm(obj=school)
    if form.validate_on_submit():
        if form.name.data != school.name and School.query.filter_by(teacher_id=current_user.id, name=form.name.data).first():
            flash('School name must be unique.', 'danger')
        else:
            school.name = form.name.data
            db.session.commit()
            flash('School updated!', 'success')
            return redirect(url_for('main.manage_schools'))
    return render_template('edit_school.html', form=form)

@main.route('/schools/delete/<int:school_id>', methods=['POST'])
@login_required
def delete_school(school_id):
    from .models import School
    school = School.query.filter_by(id=school_id, teacher_id=current_user.id).first_or_404()
    db.session.delete(school)
    db.session.commit()
    flash('School deleted!', 'info')
    return redirect(url_for('main.dashboard'))

@main.route('/classrooms', methods=['GET', 'POST'])
@login_required
def manage_classrooms():
    from .models import Classroom, School
    from .forms import ClassroomForm, ClassroomFormForSchool
    schools = School.query.filter_by(teacher_id=current_user.id).all()
    school_id = request.args.get('school_id', type=int)
    classrooms = []
    if school_id:
        classrooms = Classroom.query.filter_by(school_id=school_id).order_by(Classroom.name.asc()).all()
        form = ClassroomFormForSchool()
    else:
        classrooms = Classroom.query.join(School).filter(School.teacher_id==current_user.id).order_by(Classroom.name.asc()).all()
        form = ClassroomForm()
        form.school.choices = [(school.id, school.name) for school in schools]
    if school_id:
        if form.validate_on_submit():
            if len(classrooms) >= 20:
                flash('Maximum of 20 classrooms allowed for this school.', 'danger')
            elif any(c.name == form.name.data for c in classrooms):
                flash('Classroom name must be unique within a school.', 'danger')
            else:
                new_classroom = Classroom(name=form.name.data, school_id=school_id)
                db.session.add(new_classroom)
                db.session.commit()
                flash('Classroom added!', 'success')
                return redirect(url_for('main.manage_classrooms', school_id=school_id))
    else:
        if form.validate_on_submit():
            if len(classrooms) >= 20:
                flash('Maximum of 20 classrooms allowed.', 'danger')
            elif any(c.name == form.name.data and c.school_id == form.school.data for c in classrooms):
                flash('Classroom name must be unique within a school.', 'danger')
            else:
                new_classroom = Classroom(name=form.name.data, school_id=form.school.data)
                db.session.add(new_classroom)
                db.session.commit()
                flash('Classroom added!', 'success')
                return redirect(url_for('main.manage_classrooms', school_id=form.school.data))
    return render_template('classrooms.html', form=form, classrooms=classrooms, schools=schools, school_id=school_id, show_global_filters=True)

@main.route('/classrooms/edit/<int:classroom_id>', methods=['GET', 'POST'])
@login_required
def edit_classroom(classroom_id):
    from .models import Classroom, School
    from .forms import EditClassroomForm
    classroom = Classroom.query.join(School).filter(Classroom.id==classroom_id, School.teacher_id==current_user.id).first_or_404()
    schools = School.query.filter_by(teacher_id=current_user.id).all()
    form = EditClassroomForm(obj=classroom)
    form.school.choices = [(school.id, school.name) for school in schools]
    if form.validate_on_submit():
        if form.name.data != classroom.name or form.school.data != classroom.school_id:
            if Classroom.query.join(School).filter(School.teacher_id==current_user.id, Classroom.name==form.name.data, Classroom.school_id==form.school.data, Classroom.id!=classroom_id).first():
                flash('Classroom name must be unique within a school.', 'danger')
                return render_template('edit_classroom.html', form=form)
        classroom.name = form.name.data
        classroom.school_id = form.school.data
        db.session.commit()
        flash('Classroom updated!', 'success')
        return redirect(url_for('main.manage_classrooms'))
    form.school.data = classroom.school_id
    return render_template('edit_classroom.html', form=form)

@main.route('/classrooms/delete/<int:classroom_id>', methods=['POST'])
@login_required
def delete_classroom(classroom_id):
    from .models import Classroom, School
    classroom = Classroom.query.join(School).filter(Classroom.id==classroom_id, School.teacher_id==current_user.id).first_or_404()
    school_id = classroom.school_id
    db.session.delete(classroom)
    db.session.commit()
    flash('Classroom deleted!', 'info')
    return redirect(url_for('main.manage_classrooms', school_id=school_id))

@main.route('/students/<int:classroom_id>', methods=['GET', 'POST'])
@login_required
def manage_students(classroom_id):
    from .models import Classroom, Student, School
    from .forms import StudentForm
    classroom = Classroom.query.join(School).filter(Classroom.id==classroom_id, School.teacher_id==current_user.id).first_or_404()
    students = Student.query.filter_by(classroom_id=classroom.id).order_by(Student.last_name.asc(), Student.first_name.asc()).all()
    form = StudentForm()
    if form.validate_on_submit():
        if len(students) >= 100:
            flash('Maximum of 100 students allowed per classroom.', 'danger')
        elif any(s.first_name == form.first_name.data and s.last_name == form.last_name.data for s in students):
            flash('Student name must be unique within this classroom.', 'danger')
        else:
            new_student = Student(first_name=form.first_name.data, last_name=form.last_name.data, classroom_id=classroom.id)
            db.session.add(new_student)
            db.session.commit()
            flash('Student added!', 'success')
            return redirect(url_for('main.manage_students', classroom_id=classroom.id))
    return render_template('students.html', form=form, students=students, classroom=classroom)

@main.route('/students/delete/<int:student_id>', methods=['POST'])
@login_required
def delete_student(student_id):
    from .models import Student, Classroom, School
    student = Student.query.join(Classroom).join(School).filter(
        Student.id==student_id,
        School.teacher_id==current_user.id
    ).first_or_404()
    classroom_id = student.classroom_id
    # CASCADE delete will automatically remove all associated grades
    db.session.delete(student)
    db.session.commit()
    flash('Student and all associated grades deleted!', 'info')
    return redirect(url_for('main.manage_students', classroom_id=classroom_id))

@main.route('/api/get_test/<int:test_id>')
@login_required
def get_test(test_id):
    """Get test data for editing"""
    test = Test.query.filter_by(id=test_id, teacher_id=current_user.id).first()
    
    if not test:
        return jsonify({'error': 'Test not found'}), 404
    
    return jsonify({
        'id': test.id,
        'semester': test.semester,
        'grade': test.grade,
        'class_name': test.class_name,
        'subject': test.subject,
        'competency': test.competency,
        'test_name': test.test_name,
        'max_points': test.max_points,
        'test_date': test.test_date.strftime('%Y-%m-%d'),
        'test_weight': test.test_weight
    })

@main.route('/api/delete_test/<int:test_id>', methods=['DELETE'])
@login_required
def delete_test(test_id):
    """Delete a test"""
    try:
        test = Test.query.filter_by(id=test_id, teacher_id=current_user.id).first()
        
        if not test:
            return jsonify({'success': False, 'error': 'Test not found'}), 404
        
        db.session.delete(test)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Test deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@main.route('/api/get_test_for_grading/<int:test_id>')
@login_required
def get_test_for_grading(test_id):
    """Get test data with students for grade input"""
    from .models import Grade
    
    test = Test.query.filter_by(id=test_id, teacher_id=current_user.id).first()
    
    if not test:
        return jsonify({'error': 'Test not found'}), 404
    
    # Determine which students should be included based on teacher type and test details
    students = []
    
    # Get all classrooms for this teacher
    all_classrooms = []
    for school in School.query.filter_by(teacher_id=current_user.id).all():
        school_classrooms = Classroom.query.filter_by(school_id=school.id).all()
        all_classrooms.extend(school_classrooms)
    
    total_classrooms = len(all_classrooms)
    teacher_type = 'specialist' if total_classrooms > 1 else 'homeroom'
    
    if teacher_type == 'specialist':
        # For specialist teachers, get students from the specific class
        target_classroom = None
        current_app.logger.info(f"Looking for students - Test grade: '{test.grade}', Test class_name: '{test.class_name}'")
        current_app.logger.info(f"Available classrooms: {[(c.id, c.name) for c in all_classrooms]}")

        def _normalize_grade_label(value):
            if value is None:
                return ''
            s = str(value).strip()
            if not s:
                return ''
            lower = s.lower()
            if lower.startswith('grade '):
                s = s[6:].strip()
            return s

        normalized_test_grade = _normalize_grade_label(test.grade)
        class_name_candidates = []
        
        for classroom in all_classrooms:
            if '(' in classroom.name and ')' in classroom.name:
                parts = classroom.name.split(' (')
                classroom_name = parts[0]
                grade = parts[1].rstrip(')')
            else:
                classroom_name = classroom.name
                grade = extract_grade_from_classroom_name(classroom.name)
            
            current_app.logger.info(f"Checking classroom: name='{classroom_name}', grade='{grade}' against test grade='{test.grade}', class_name='{test.class_name}'")
            
            # Primary match is class name. Grade label is used only as a tie-breaker since
            # specialist grade labels can be teacher-defined (e.g., "5" vs "Grade 5").
            class_name_matches = (
                test.class_name == classroom_name or
                test.class_name == classroom.name
            )

            if not class_name_matches:
                continue

            normalized_classroom_grade = _normalize_grade_label(grade)
            grade_matches = (not normalized_test_grade) or (normalized_test_grade == normalized_classroom_grade)

            class_name_candidates.append((classroom, grade_matches, normalized_classroom_grade))

        if class_name_candidates:
            # Prefer the candidate whose grade matches (if grade is supplied). Otherwise,
            # fall back to the first class-name match.
            matching_grade_candidates = [c for c in class_name_candidates if c[1]]
            if matching_grade_candidates:
                target_classroom = matching_grade_candidates[0][0]
            else:
                target_classroom = class_name_candidates[0][0]

            current_app.logger.info(
                "Selected classroom for grading: "
                f"test_id={test_id} test_class='{test.class_name}' test_grade='{test.grade}' "
                f"-> classroom_id={target_classroom.id} classroom_name='{target_classroom.name}'"
            )
        
        if target_classroom:
            students = Student.query.filter_by(classroom_id=target_classroom.id).order_by(Student.last_name, Student.first_name).all()
            current_app.logger.info(f"Found {len(students)} students in classroom {target_classroom.id}")
        else:
            current_app.logger.warning(f"No matching classroom found for test {test_id}")
    else:
        # For homeroom teachers, get all students from their single classroom
        if all_classrooms:
            students = Student.query.filter_by(classroom_id=all_classrooms[0].id).order_by(Student.last_name, Student.first_name).all()
    
    # Get existing grades for these students
    existing_grades = {}
    if students:
        student_ids = [s.id for s in students]
        grades = Grade.query.filter(
            Grade.test_id == test_id,
            Grade.student_id.in_(student_ids)
        ).all()
        
        for grade in grades:
            existing_grades[grade.student_id] = {
                'grade': grade.grade,
                'absent': grade.absent
            }
    
    # Format student data with existing grades
    students_data = []
    for student in students:
        grade_data = existing_grades.get(student.id, {})
        students_data.append({
            'id': student.id,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'grade': grade_data.get('grade') if isinstance(grade_data, dict) else grade_data,
            'absent': grade_data.get('absent', False) if isinstance(grade_data, dict) else False
        })
    
    return jsonify({
        'test': {
            'id': test.id,
            'semester': test.semester,
            'grade': test.grade,
            'class_name': test.class_name,
            'subject': test.subject,
            'competency': test.competency,
            'test_name': test.test_name,
            'max_points': test.max_points,
            'test_date': test.test_date.strftime('%Y-%m-%d'),
            'test_weight': test.test_weight
        },
        'students': students_data
    })

@main.route('/api/save_grades/<int:test_id>', methods=['POST'])
@login_required
def save_grades(test_id):
    """Save grades for a test"""
    from .models import Grade
    
    try:
        test = Test.query.filter_by(id=test_id, teacher_id=current_user.id).first()
        
        if not test:
            return jsonify({'success': False, 'error': 'Test not found'}), 404
        
        grades_data = request.json.get('grades', [])
        
        for grade_data in grades_data:
            student_id = grade_data['student_id']
            grade_value = grade_data['grade']
            absent_value = grade_data.get('absent', False)
            
            current_app.logger.info(f"Processing grade for student {student_id}: grade={grade_value}, absent={absent_value}")
            
            # Check if grade already exists
            existing_grade = Grade.query.filter_by(
                test_id=test_id,
                student_id=student_id
            ).first()
            
            # Only save if there's actual data (grade exists OR student is absent)
            # Skip students with no grade and not absent (empty records)
            if grade_value is not None or absent_value:
                if existing_grade:
                    # Update existing grade - only update current values, preserve original
                    existing_grade.grade = grade_value
                    existing_grade.absent = absent_value
                    existing_grade.updated_at = datetime.utcnow()
                    current_app.logger.info(f"Updated grade for student {student_id}")
                else:
                    # Create new grade - set both current and original values
                    new_grade = Grade(
                        test_id=test_id,
                        student_id=student_id,
                        grade=grade_value,
                        absent=absent_value,
                        original_grade=grade_value,
                        original_absent=absent_value
                    )
                    db.session.add(new_grade)
                    current_app.logger.info(f"Created new grade for student {student_id}")
            else:
                # If there's an existing grade record but now it's empty, delete it
                if existing_grade:
                    db.session.delete(existing_grade)
                    current_app.logger.info(f"Deleted empty grade record for student {student_id}")
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Grades saved successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@main.route('/api/get_grade_matrix')
@login_required
def get_grade_matrix():
    """Get grade matrix data for review grades page"""
    from .models import Test, Student, Grade, Classroom, School, SetupWizardData
    import json
    
    # Get filter parameters
    semester = request.args.get('semester', '')
    class_name = request.args.get('class_name', '')
    subject = request.args.get('subject', '')
    competency = request.args.get('competency', '')
    
    try:
        # Debug logging
        print(f"DEBUG: get_grade_matrix called with semester='{semester}', class_name='{class_name}', subject='{subject}', competency='{competency}'")
        
        # Build test query with filters
        test_query = Test.query.filter_by(teacher_id=current_user.id)
        
        if semester:
            test_query = test_query.filter(Test.semester == semester)
        if class_name:
            test_query = test_query.filter(Test.class_name == class_name)
        if subject:
            test_query = test_query.filter(Test.subject == subject)
        if competency:
            test_query = test_query.filter(Test.competency == competency)
        
        tests = test_query.order_by(Test.test_date).all()
        print(f"DEBUG: Found {len(tests)} tests")
        
        if not tests:
            print("DEBUG: No tests found, returning empty response")
            return jsonify({'tests': [], 'students': [], 'grades': {}})
        
        # Get students from the relevant classroom(s)
        students = []
        if class_name:
            # For specialist teachers - specific class
            # Find the classroom that matches the class name
            for school in School.query.filter_by(teacher_id=current_user.id).all():
                for classroom in Classroom.query.filter_by(school_id=school.id).all():
                    print(f"DEBUG: Checking classroom '{classroom.name}' against class_name '{class_name}'")
                    # Extract class name from classroom name (format: "ClassName (Grade X)")
                    classroom_class_name = classroom.name.split(' (')[0] if ' (' in classroom.name else classroom.name
                    if classroom_class_name == class_name:
                        classroom_students = Student.query.filter_by(classroom_id=classroom.id).all()
                        students.extend(classroom_students)
                        print(f"DEBUG: Found matching classroom, added {len(classroom_students)} students")
        else:
            # Get all students from all classrooms for this teacher
            for school in School.query.filter_by(teacher_id=current_user.id).all():
                for classroom in Classroom.query.filter_by(school_id=school.id).all():
                    classroom_students = Student.query.filter_by(classroom_id=classroom.id).all()
                    students.extend(classroom_students)
        
        print(f"DEBUG: Total students found: {len(students)}")
        
        # Get all grades for these tests and students
        test_ids = [test.id for test in tests]
        student_ids = [student.id for student in students]
        
        grades_query = Grade.query.filter(
            Grade.test_id.in_(test_ids),
            Grade.student_id.in_(student_ids)
        ).all()
        
        # Organize grades by student_id and test_id
        grades_matrix = {}
        for grade_record in grades_query:
            if grade_record.student_id not in grades_matrix:
                grades_matrix[grade_record.student_id] = {}
            grades_matrix[grade_record.student_id][grade_record.test_id] = grade_record.grade
        
        # Format response data
        tests_data = []
        for test in tests:
            tests_data.append({
                'id': test.id,
                'test_name': test.test_name,
                'max_points': test.max_points,
                'test_weight': test.test_weight,
                'test_date': test.test_date.strftime('%Y-%m-%d'),
                'competency': test.competency
            })
        
        students_data = []
        for student in students:
            students_data.append({
                'id': student.id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'full_name': f"{student.first_name} {student.last_name}"
            })
        
        # Get competency weights from Setup Wizard data
        wizard_data = SetupWizardData.query.filter_by(teacher_id=current_user.id).first()
        competency_weights = {}
        if wizard_data and wizard_data.weights and wizard_data.competencies:
            try:
                weights_data = json.loads(wizard_data.weights)
                competencies_list = json.loads(wizard_data.competencies)
                
                # Extract weights for the current grade/semester if available
                # The structure is: {grade: {semester: {competency_index: weight}}}
                # For now, use the first available grade/semester combination
                for grade_key in weights_data:
                    for semester_key in weights_data[grade_key]:
                        semester_weights = weights_data[grade_key][semester_key]
                        for comp_index, weight in semester_weights.items():
                            comp_index_int = int(comp_index)
                            if comp_index_int < len(competencies_list):
                                competency_name = competencies_list[comp_index_int]
                                competency_weights[competency_name] = int(weight)
                        break  # Use first semester found
                    break  # Use first grade found
            except Exception as e:
                print(f"Error parsing competency weights: {e}")
                competency_weights = {}
        
        return jsonify({
            'tests': tests_data,
            'students': students_data,
            'grades': grades_matrix,
            'competency_weights': competency_weights
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main.route('/api/save_grade_updates', methods=['POST'])
@login_required
def save_grade_updates():
    """Save grade updates from review grades page"""
    from .models import Grade, Test
    
    try:
        data = request.get_json()
        updates = data.get('updates', [])
        
        for update in updates:
            student_id = update['student_id']
            test_id = update['test_id']
            grade_value = update['grade']
            
            # Verify the test belongs to the current teacher
            test = Test.query.filter_by(id=test_id, teacher_id=current_user.id).first()
            if not test:
                return jsonify({'success': False, 'error': f'Test {test_id} not found'}), 404
            
            # Find existing grade record
            existing_grade = Grade.query.filter_by(test_id=test_id, student_id=student_id).first()
            
            if existing_grade:
                if grade_value is not None:
                    existing_grade.grade = grade_value
                    existing_grade.updated_at = datetime.utcnow()
                else:
                    # Delete grade if value is None/null
                    db.session.delete(existing_grade)
            else:
                if grade_value is not None:
                    # Create new grade record
                    new_grade = Grade(
                        test_id=test_id,
                        student_id=student_id,
                        grade=grade_value
                    )
                    db.session.add(new_grade)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Grades updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@main.route('/admin/run_migration_cascade_delete', methods=['GET'])
@login_required
def run_migration_cascade_delete():
    """Temporary endpoint to run CASCADE delete migration on production.
    DELETE THIS ENDPOINT AFTER RUNNING ONCE!"""
    from sqlalchemy import text, inspect
    
    try:
        # Check if CASCADE delete is already set up by trying to query the constraint
        inspector = inspect(db.engine)
        
        # Try to check if the constraint already has CASCADE
        # We'll just run the migration - it's idempotent with DROP IF EXISTS
        with db.engine.connect() as conn:
            # Drop the existing foreign key constraint
            conn.execute(text("""
                ALTER TABLE grade 
                DROP CONSTRAINT IF EXISTS grade_student_id_fkey;
            """))
            
            # Add the foreign key constraint with CASCADE delete
            conn.execute(text("""
                ALTER TABLE grade 
                ADD CONSTRAINT grade_student_id_fkey 
                FOREIGN KEY (student_id) 
                REFERENCES student(id) 
                ON DELETE CASCADE;
            """))
            
            conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'CASCADE delete migration completed successfully! Grades will now be automatically deleted when a student is deleted.',
            'action': 'completed'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        })
